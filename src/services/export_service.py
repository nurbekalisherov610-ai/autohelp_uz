import io
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from src.db.models.order import Order
from src.db.models.user import User

class ExportImportService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def export_orders(self) -> bytes:
        orders = (await self.session.scalars(select(Order).order_by(Order.created_at.desc()))).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(["ID", "Client ID", "Phone", "Issue", "Status", "Amount", "Rating", "Feedback", "Shortcomings", "Created At", "Completed At"])
        for o in orders:
            ws.append([
                o.id,
                o.client_id,
                o.phone,
                o.issue_label,
                o.status.name,
                float(o.final_amount) if o.final_amount else 0,
                o.rating,
                o.feedback_text,
                o.shortcomings,
                str(o.created_at),
                str(o.completed_at) if o.completed_at else ""
            ])
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    async def export_users(self) -> bytes:
        users = (await self.session.scalars(select(User).order_by(User.created_at.desc()))).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["ID", "Telegram ID", "Full Name", "Phone", "Language", "Is Master", "Is Blocked", "Created At"])
        for u in users:
            ws.append([
                u.id,
                u.telegram_id,
                u.full_name,
                u.phone,
                u.language,
                u.is_master,
                u.is_blocked,
                str(u.created_at)
            ])
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()

    async def import_users(self, excel_bytes: bytes) -> tuple[int, int]:
        stream = io.BytesIO(excel_bytes)
        wb = load_workbook(stream)
        ws = wb.active
        
        imported = 0
        updated = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            
            telegram_id = int(row[1])
            full_name = str(row[2]) if row[2] else None
            phone = str(row[3]) if row[3] else None
            language = str(row[4]) if row[4] else None
            is_master = bool(row[5])
            is_blocked = bool(row[6])
            
            user = await self.session.scalar(select(User).where(User.telegram_id == telegram_id))
            if user:
                user.full_name = full_name
                user.phone = phone
                user.language = language
                user.is_master = is_master
                user.is_blocked = is_blocked
                updated += 1
            else:
                user = User(
                    telegram_id=telegram_id,
                    full_name=full_name,
                    phone=phone,
                    language=language,
                    is_master=is_master,
                    is_blocked=is_blocked
                )
                self.session.add(user)
                imported += 1
        
        await self.session.commit()
        return imported, updated
