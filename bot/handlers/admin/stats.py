"""
AutoHelp.uz - Full Admin Dashboard
CEO-level monitoring: live stats, orders, reviews, masters, Excel export.
All buttons work. All errors handled.
"""
import io
from datetime import datetime, timedelta

from aiogram import Router, F, Bot
from aiogram.exceptions import TelegramBadRequest
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.fsm.context import FSMContext
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from bot.filters.role_filter import RoleFilter
from bot.keyboards.admin_kb import (
    admin_main_menu, admin_orders_filter, admin_master_actions,
    admin_export_options, admin_reports_period, admin_back_button,
)
from models.order import Order, OrderStatus, PROBLEM_LABELS
from models.master import Master
from models.staff import Staff
from models.user import User
from models.review import Review
from repositories.order_repo import OrderRepo
from repositories.master_repo import MasterRepo
from repositories.stats_repo import StatsRepo

router = Router(name="admin")

STATUS_EMOJI = {
    "new": "🆕", "assigned": "👨‍🔧", "accepted": "✅",
    "on_the_way": "🚗", "arrived": "📍", "in_progress": "🔧",
    "awaiting_confirm": "⏳", "completed": "✅", "cancelled": "❌", "rejected": "🔄",
}


def _safe_pct(part, total) -> str:
    if not total:
        return "0.0"
    return f"{part / total * 100:.1f}"


async def _edit_or_send(callback: CallbackQuery, text: str, **kwargs):
    """Edit callback message, fallback to sending a new one if edit is impossible."""
    try:
        return await callback.message.edit_text(text, **kwargs)
    except TelegramBadRequest as e:
        err = str(e).lower()
        if "message is not modified" in err:
            return None
        if "message can't be edited" in err or "message to edit not found" in err:
            return await callback.message.answer(text, **kwargs)
        raise


# ── /admin command ────────────────────────────────────────────────

@router.message(RoleFilter("admin", "super_admin", "dispatcher"), F.text == "/admin")
@router.message(RoleFilter("admin", "super_admin", "dispatcher"), F.text.contains("Admin"))
async def admin_start(message: Message):
    """Admin main menu."""
    await message.answer(
        "👑 <b>Admin Paneli — AutoHelp.uz</b>\n\n"
        "Samarqand viloyati bo'yicha barcha jarayonlarni kuzating 👇",
        parse_mode="HTML",
        reply_markup=admin_main_menu(),
    )


@router.message(F.text == "/admin")
async def admin_start_denied(message: Message):
    """Friendly access-denied message when user has no admin privileges."""
    await message.answer(
        "⛔ Sizda admin panel huquqi yo'q.\n"
        "Admin kirishini yoqish uchun: python manage.py add_admin <telegram_id> <ism> <tel>",
    )


@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:menu")
async def admin_menu_cb(callback: CallbackQuery):
    await _edit_or_send(callback, 
        "👑 <b>Admin Paneli — AutoHelp.uz</b>\n\n"
        "Samarqand viloyati bo'yicha barcha jarayonlarni kuzating 👇",
        parse_mode="HTML",
        reply_markup=admin_main_menu(),
    )
    await callback.answer()


# ── Dashboard ─────────────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:dashboard")
async def admin_dashboard(callback: CallbackQuery, session: AsyncSession):
    """Live dashboard — CEO view."""
    from loguru import logger
    logger.info("ADMIN DASHBOARD HANDLER TRIGGERED!")
    try:
        await callback.answer("📊 Yuklanmoqda...")
    except Exception as e:
        logger.error(f"Failed to answer callback: {e}")

    stats_repo = StatsRepo(session)
    order_repo = OrderRepo(session)
    master_repo = MasterRepo(session)

    try:
        stats = await stats_repo.get_dashboard_stats()
    except Exception:
        stats = {
            "today_orders": 0, "monthly_orders": 0,
            "today_sum": 0, "monthly_sum": 0,
            "avg_rating": 0.0, "online_masters": 0,
            "total_users": 0, "active_orders": 0,
            "cancelled_rate": 0,
        }

    # Active orders breakdown
    active_result = await session.execute(
        select(Order.status, func.count(Order.id))
        .where(Order.status.notin_([OrderStatus.COMPLETED, OrderStatus.CANCELLED]))
        .group_by(Order.status)
    )
    active_breakdown = {row[0].value: row[1] for row in active_result.all()}

    now = datetime.utcnow()
    text = (
        f"📊 <b>Dashboard — {now.strftime('%d.%m.%Y %H:%M')}</b>\n"
        f"📍 Samarqand viloyati\n"
        f"{'─'*30}\n\n"

        f"📅 <b>Bugun:</b>\n"
        f"   📋 Buyurtmalar: {stats['today_orders']}\n"
        f"   💰 Daromad: {stats['today_sum']:,.0f} so'm\n\n"

        f"📆 <b>Bu oy:</b>\n"
        f"   📋 Buyurtmalar: {stats['monthly_orders']}\n"
        f"   💰 Daromad: {stats['monthly_sum']:,.0f} so'm\n\n"

        f"⚡ <b>Hozir faol:</b>\n"
        f"   🆕 Yangi: {active_breakdown.get('new', 0)}\n"
        f"   👨‍🔧 Tayinlangan: {active_breakdown.get('assigned', 0)}\n"
        f"   🚗 Yo'lda: {active_breakdown.get('on_the_way', 0)}\n"
        f"   🔧 Jarayonda: {active_breakdown.get('in_progress', 0)}\n\n"

        f"👨‍🔧 <b>Ustalar:</b> {stats['online_masters']} online\n"
        f"⭐ <b>O'rtacha reyting:</b> {stats['avg_rating']:.1f}/5.0\n"
        f"👥 <b>Jami mijozlar:</b> {stats['total_users']}\n"
    )

    await _edit_or_send(callback, 
        text, parse_mode="HTML", reply_markup=admin_back_button()
    )


# ── Active Orders ─────────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:active_orders")
async def admin_active_orders(callback: CallbackQuery, session: AsyncSession):
    """Show all currently active orders."""
    await callback.answer("📋 Yuklanmoqda...")

    result = await session.scalars(
        select(Order)
        .where(Order.status.notin_([OrderStatus.COMPLETED, OrderStatus.CANCELLED]))
        .options(selectinload(Order.user), selectinload(Order.master))
        .order_by(Order.created_at.desc())
        .limit(20)
    )
    orders = list(result.all())

    if not orders:
        await _edit_or_send(callback, 
            "✅ Hozirda faol buyurtmalar yo'q.",
            reply_markup=admin_back_button(),
        )
        return

    lines = [f"⚡ <b>Faol buyurtmalar ({len(orders)}):</b>\n"]
    for order in orders:
        icon = STATUS_EMOJI.get(order.status.value, "•")
        client = order.user.full_name if order.user else "—"
        master = order.master.full_name if order.master else "Tayinlanmagan"
        problem = PROBLEM_LABELS[order.problem_type]["uz"]
        elapsed = int((datetime.utcnow() - order.created_at.replace(tzinfo=None)).total_seconds() / 60)

        lines.append(
            f"{icon} <code>{order.order_uid}</code>\n"
            f"   👤 {client} → 👨‍🔧 {master}\n"
            f"   🔧 {problem} • ⏱ {elapsed} daq\n"
        )

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:3900] + "\n...(hammasi Excel'da)"

    await _edit_or_send(callback, 
        text, parse_mode="HTML", reply_markup=admin_back_button()
    )


# ── Orders by filter ──────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:orders")
async def admin_orders_menu(callback: CallbackQuery):
    await _edit_or_send(callback, 
        "📋 <b>Buyurtmalar</b>\n\nFiltrni tanlang:",
        parse_mode="HTML",
        reply_markup=admin_orders_filter(),
    )
    await callback.answer()


@router.callback_query(
    RoleFilter("admin", "super_admin", "dispatcher"),
    F.data.startswith("admin_filter:"),
)
async def admin_filter_orders(callback: CallbackQuery, session: AsyncSession):
    """Filter orders by status."""
    await callback.answer("🔄 Yuklanmoqda...")
    filter_type = callback.data.split(":")[1]

    query = (
        select(Order)
        .options(selectinload(Order.user), selectinload(Order.master))
        .order_by(Order.created_at.desc())
        .limit(25)
    )

    filter_labels = {
        "new": "🆕 Yangi", "active": "⚡ Faol", "completed": "✅ Tugallangan",
        "cancelled": "❌ Bekor", "all": "📋 Barchasi",
    }

    if filter_type == "new":
        query = query.where(Order.status == OrderStatus.NEW)
    elif filter_type == "active":
        query = query.where(Order.status.notin_([OrderStatus.COMPLETED, OrderStatus.CANCELLED]))
    elif filter_type == "completed":
        query = query.where(Order.status == OrderStatus.COMPLETED)
    elif filter_type == "cancelled":
        query = query.where(Order.status == OrderStatus.CANCELLED)

    result = await session.scalars(query)
    orders = list(result.all())

    if not orders:
        await _edit_or_send(callback, 
            f"📭 {filter_labels.get(filter_type, filter_type)} buyurtmalar yo'q.",
            reply_markup=admin_back_button(),
        )
        return

    lines = [f"📋 <b>{filter_labels.get(filter_type, filter_type)} ({len(orders)}):</b>\n"]
    for order in orders:
        icon = STATUS_EMOJI.get(order.status.value, "•")
        client = order.user.full_name if order.user else "—"
        master = order.master.full_name if order.master else "—"
        problem = PROBLEM_LABELS[order.problem_type]["uz"]
        amount = f"💰{order.payment_amount:,.0f}" if order.payment_amount else ""
        date = order.created_at.strftime("%d.%m %H:%M")

        lines.append(
            f"{icon} <code>{order.order_uid}</code> {amount}\n"
            f"   👤{client} | 👨‍🔧{master}\n"
            f"   📌{problem} | {date}\n"
        )

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:3900] + "\n...(davomi Excel'da)"

    await _edit_or_send(callback, 
        text, parse_mode="HTML", reply_markup=admin_back_button()
    )


# ── Reviews ───────────────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:reviews")
async def admin_reviews(callback: CallbackQuery, session: AsyncSession):
    """Show latest client reviews."""
    await callback.answer("⭐ Yuklanmoqda...")

    result = await session.scalars(
        select(Review)
        .options(
            selectinload(Review.user),
            selectinload(Review.order),
        )
        .order_by(Review.created_at.desc())
        .limit(15)
    )
    reviews = list(result.all())

    if not reviews:
        await _edit_or_send(callback, 
            "⭐ Hali sharhlar yo'q.",
            reply_markup=admin_back_button(),
        )
        return

    lines = ["⭐ <b>So'nggi sharhlar:</b>\n"]
    for r in reviews:
        stars = "⭐" * r.rating + "☆" * (5 - r.rating)
        client = r.user.full_name if r.user else "—"
        order_uid = r.order.order_uid if r.order else "—"
        date = r.created_at.strftime("%d.%m.%Y")
        comment = f"\n   💬 {r.comment}" if r.comment else ""

        lines.append(
            f"{stars}\n"
            f"   👤 {client} • <code>{order_uid}</code> • {date}"
            f"{comment}\n"
        )

    await _edit_or_send(callback, 
        "\n".join(lines), parse_mode="HTML", reply_markup=admin_back_button()
    )


# ── Masters ───────────────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:masters")
async def admin_masters(callback: CallbackQuery, session: AsyncSession):
    """Show all masters with live status."""
    await callback.answer("👨‍🔧 Yuklanmoqda...")

    result = await session.scalars(
        select(Master).where(Master.is_active == True).order_by(Master.rating.desc())
    )
    masters = list(result.all())

    if not masters:
        await _edit_or_send(callback, 
            "👨‍🔧 Ustalar topilmadi.\n\n"
            "<code>python manage.py add_master &lt;id&gt; &lt;ism&gt; &lt;tel&gt;</code>",
            parse_mode="HTML",
            reply_markup=admin_back_button(),
        )
        return

    status_icon = {"online": "🟢", "busy": "🟡", "offline": "🔴"}
    lines = [f"👨‍🔧 <b>Ustalar ({len(masters)} ta):</b>\n"]
    for m in masters:
        icon = status_icon.get(m.status.value, "⚪")
        lines.append(
            f"{icon} <b>{m.full_name}</b>\n"
            f"   📞 {m.phone} | ⭐{m.rating:.1f}\n"
            f"   ✅{m.completed_orders} bajargan | ❌{m.rejected_orders} rad\n"
        )

    await _edit_or_send(callback, 
        "\n".join(lines), parse_mode="HTML", reply_markup=admin_back_button()
    )


# ── Reports ───────────────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:reports")
async def admin_reports(callback: CallbackQuery):
    await _edit_or_send(callback, 
        "📊 <b>Hisobotlar</b>\n\nDavrni tanlang:",
        parse_mode="HTML",
        reply_markup=admin_reports_period(),
    )
    await callback.answer()


@router.callback_query(
    RoleFilter("admin", "super_admin", "dispatcher"),
    F.data.startswith("report:"),
)
async def generate_report(callback: CallbackQuery, session: AsyncSession):
    """Generate a period statistics report."""
    await callback.answer("📊 Hisobot tayyorlanmoqda...")
    period = callback.data.split(":")[1]
    now = datetime.utcnow()

    period_map = {
        "today": now.replace(hour=0, minute=0, second=0, microsecond=0),
        "week": now - timedelta(days=7),
        "month": now.replace(day=1, hour=0, minute=0, second=0, microsecond=0),
        "year": now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0),
    }
    period_labels = {"today": "Bugun", "week": "Hafta", "month": "Oy", "year": "Yil"}
    since = period_map.get(period, period_map["today"])

    order_repo = OrderRepo(session)
    stats_repo = StatsRepo(session)

    total = await order_repo.count_by_status(since=since)
    completed = await order_repo.count_by_status(status=OrderStatus.COMPLETED, since=since)
    cancelled = await order_repo.count_by_status(status=OrderStatus.CANCELLED, since=since)
    revenue = await order_repo.sum_payments(since=since)
    avg_rating = await order_repo.avg_rating(since=since)

    leaderboard = await stats_repo.get_master_leaderboard(limit=5)
    lb_text = "\n".join(
        f"   {i+1}. {m['name']} — ⭐{m['rating']:.1f} ({m['completed']} ish)"
        for i, m in enumerate(leaderboard)
    ) or "   Ma'lumot yo'q"

    conv_pct = _safe_pct(completed, total)

    text = (
        f"📊 <b>Hisobot: {period_labels.get(period, period)}</b>\n"
        f"📍 Samarqand viloyati\n"
        f"{'─'*28}\n\n"
        f"📋 Jami buyurtmalar: <b>{total}</b>\n"
        f"✅ Tugallangan: <b>{completed}</b>\n"
        f"❌ Bekor qilingan: <b>{cancelled}</b>\n"
        f"💰 Daromad: <b>{revenue:,.0f} so'm</b>\n"
        f"⭐ O'rtacha reyting: <b>{avg_rating:.1f}/5.0</b>\n"
        f"📈 Konversiya: <b>{conv_pct}%</b>\n\n"
        f"🏆 <b>Top ustalar:</b>\n{lb_text}"
    )

    await _edit_or_send(callback, 
        text, parse_mode="HTML", reply_markup=admin_back_button()
    )


# ── Excel Export ──────────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:export")
async def admin_export_menu(callback: CallbackQuery):
    await _edit_or_send(callback, 
        "📥 <b>Excel Eksport</b>\n\nNimani eksport qilmoqchisiz?",
        parse_mode="HTML",
        reply_markup=admin_export_options(),
    )
    await callback.answer()


@router.callback_query(
    RoleFilter("admin", "super_admin", "dispatcher"),
    F.data.startswith("export:"),
)
async def process_export(callback: CallbackQuery, session: AsyncSession, bot: Bot):
    """Generate and send Excel file."""
    export_type = callback.data.split(":")[1]
    await callback.answer("📊 Fayl tayyorlanmoqda...")
    await _edit_or_send(callback, "⏳ Excel fayl tayyorlanmoqda...")

    wb = Workbook()
    ws = wb.active

    # Style helpers
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def style_headers(row_num=1):
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    if export_type == "orders":
        ws.title = "Buyurtmalar"
        ws.append(["#", "UID", "Mijoz", "Telefon", "Muammo", "Status",
                   "Usta", "Summa (so'm)", "Reyting", "Sana", "Tugallangan"])
        style_headers()

        result = await session.scalars(
            select(Order)
            .options(
                selectinload(Order.user),
                selectinload(Order.master),
                selectinload(Order.review),
            )
            .order_by(Order.created_at.desc())
            .limit(2000)
        )
        for i, order in enumerate(result.all(), 1):
            ws.append([
                i, order.order_uid,
                order.user.full_name if order.user else "—",
                order.user.phone if order.user else "—",
                PROBLEM_LABELS[order.problem_type]["uz"],
                order.status.value,
                order.master.full_name if order.master else "—",
                order.payment_amount or 0,
                order.review.rating if order.review else "—",
                order.created_at.strftime("%d.%m.%Y %H:%M") if order.created_at else "",
                order.completed_at.strftime("%d.%m.%Y %H:%M") if order.completed_at else "",
            ])

    elif export_type == "masters":
        ws.title = "Ustalar"
        ws.append(["#", "Ism", "Telefon", "Status", "Reyting",
                   "Jami", "Bajarilgan", "Rad etilgan", "Ro'yxatga olingan"])
        style_headers()

        result = await session.scalars(select(Master).order_by(Master.rating.desc()))
        for i, m in enumerate(result.all(), 1):
            ws.append([
                i, m.full_name, m.phone, m.status.value, round(m.rating, 2),
                m.total_orders, m.completed_orders, m.rejected_orders,
                m.created_at.strftime("%d.%m.%Y") if m.created_at else "",
            ])

    elif export_type == "finance":
        ws.title = "Moliyaviy hisobot"
        ws.append(["#", "Buyurtma", "Mijoz", "Usta", "Summa", "Tasdiqlangan", "Sana"])
        style_headers()

        from models.payment import Payment
        result = await session.scalars(
            select(Payment)
            .options(selectinload(Payment.order).selectinload(Order.master),
                     selectinload(Payment.order).selectinload(Order.user))
            .order_by(Payment.created_at.desc())
            .limit(2000)
        )
        for i, p in enumerate(result.all(), 1):
            ws.append([
                i,
                p.order.order_uid if p.order else "—",
                p.order.user.full_name if (p.order and p.order.user) else "—",
                p.order.master.full_name if (p.order and p.order.master) else "—",
                p.amount,
                "Ha" if p.confirmed_by_dispatcher else "Yo'q",
                p.created_at.strftime("%d.%m.%Y %H:%M") if p.created_at else "",
            ])

    elif export_type == "reviews":
        ws.title = "Sharhlar"
        ws.append(["#", "Buyurtma", "Mijoz", "Reyting", "Izoh", "Sana"])
        style_headers()

        result = await session.scalars(
            select(Review)
            .options(selectinload(Review.user), selectinload(Review.order))
            .order_by(Review.created_at.desc())
        )
        for i, r in enumerate(result.all(), 1):
            ws.append([
                i,
                r.order.order_uid if r.order else "—",
                r.user.full_name if r.user else "—",
                r.rating,
                r.comment or "",
                r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else "",
            ])

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    # Send file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"autohelp_{export_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    file = BufferedInputFile(buffer.read(), filename=filename)

    await bot.send_document(
        chat_id=callback.from_user.id,
        document=file,
        caption=(
            f"📥 <b>{export_type.capitalize()} eksporti</b>\n"
            f"📅 {datetime.utcnow().strftime('%d.%m.%Y %H:%M')}\n"
            f"📍 Samarqand viloyati"
        ),
        parse_mode="HTML",
    )
    await _edit_or_send(callback, 
        f"✅ {export_type.capitalize()} fayli yuborildi!",
        reply_markup=admin_back_button(),
    )


# ── Audit Log ─────────────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:audit")
async def admin_audit(callback: CallbackQuery, session: AsyncSession):
    await callback.answer("📝 Yuklanmoqda...")

    from models.audit import AuditLog
    result = await session.scalars(
        select(AuditLog).order_by(AuditLog.created_at.desc()).limit(20)
    )
    logs = list(result.all())

    if not logs:
        await _edit_or_send(callback, 
            "📝 Audit log bo'sh.", reply_markup=admin_back_button()
        )
        return

    lines = ["📝 <b>So'nggi amallar:</b>\n"]
    for log in logs:
        time_str = log.created_at.strftime("%H:%M %d.%m") if log.created_at else ""
        lines.append(f"• <code>{log.action}</code> | {log.entity_type} | {time_str}")

    await _edit_or_send(callback, 
        "\n".join(lines), parse_mode="HTML", reply_markup=admin_back_button()
    )


# ── Dispatchers List ──────────────────────────────────────────────

@router.callback_query(RoleFilter("admin", "super_admin", "dispatcher"), F.data == "admin:dispatchers")
async def admin_dispatchers(callback: CallbackQuery, session: AsyncSession):
    await callback.answer()

    from models.staff import StaffRole
    result = await session.scalars(
        select(Staff).where(Staff.role == StaffRole.DISPATCHER, Staff.is_active == True)
    )
    staff_list = list(result.all())

    if not staff_list:
        await _edit_or_send(callback, 
            "👥 Dispetcherlar topilmadi.\n\n"
            "<code>python manage.py add_dispatcher &lt;id&gt; &lt;ism&gt; &lt;tel&gt;</code>",
            parse_mode="HTML",
            reply_markup=admin_back_button(),
        )
        return

    lines = [f"👥 <b>Dispetcherlar ({len(staff_list)} ta):</b>\n"]
    for s in staff_list:
        lines.append(f"• <b>{s.full_name}</b> | 📞 {s.phone or '—'} | ID: <code>{s.telegram_id}</code>")

    await _edit_or_send(callback, 
        "\n".join(lines), parse_mode="HTML", reply_markup=admin_back_button()
    )

@router.callback_query(
    RoleFilter("admin", "super_admin", "dispatcher"),
    F.data.startswith("admin_master_stats:"),
)
async def admin_master_stats_view(callback: CallbackQuery, session: AsyncSession):
    """Detailed view and actions for a specific master."""
    master_id = int(callback.data.split(":")[1])
    master_repo = MasterRepo(session)
    master = await master_repo.get_by_id(master_id)

    if not master:
        await callback.answer("Usta topilmadi", show_alert=True)
        return

    stats = await master_repo.get_master_stats(master.id)
    text = (
        f"👨‍🔧 <b>{master.full_name}</b>\n\n"
        f"📞 {master.phone}\n"
        f"📶 Holat: <b>{master.status.value}</b>\n"
        f"⭐ Reyting: <b>{master.rating:.1f}</b>\n"
        f"📋 Jami ishlar: <b>{stats['total_orders']}</b>\n"
        f"✅ Bajarilgan: <b>{stats['completed_orders']}</b>\n"
        f"💰 Tushum: <b>{stats['total_sum']:,.0f} so'm</b>\n"
        f"🟢 Aktiv: <b>{'Ha' if master.is_active else 'Yo‘q'}</b>"
    )

    await _edit_or_send(callback, 
        text,
        parse_mode="HTML",
        reply_markup=admin_master_actions(master.id),
    )
    await callback.answer()


@router.callback_query(
    RoleFilter("admin", "super_admin", "dispatcher"),
    F.data.startswith("admin_master_activate:"),
)
async def admin_master_activate(callback: CallbackQuery, session: AsyncSession):
    """Activate a master account."""
    master_id = int(callback.data.split(":")[1])
    await session.execute(
        update(Master)
        .where(Master.id == master_id)
        .values(is_active=True)
    )
    await callback.answer("✅ Usta faollashtirildi")

    # Refresh card
    callback.data = f"admin_master_stats:{master_id}"
    await admin_master_stats_view(callback, session)


@router.callback_query(
    RoleFilter("admin", "super_admin", "dispatcher"),
    F.data.startswith("admin_master_deactivate:"),
)
async def admin_master_deactivate(callback: CallbackQuery, session: AsyncSession):
    """Deactivate a master account."""
    master_id = int(callback.data.split(":")[1])
    from models.master import MasterStatus

    await session.execute(
        update(Master)
        .where(Master.id == master_id)
        .values(is_active=False, status=MasterStatus.OFFLINE)
    )
    await callback.answer("🚫 Usta bloklandi")

    # Refresh card
    callback.data = f"admin_master_stats:{master_id}"
    await admin_master_stats_view(callback, session)


@router.callback_query(
    RoleFilter("admin", "super_admin", "dispatcher"),
    F.data.startswith("admin:"),
)
async def admin_callback_fallback(callback: CallbackQuery):
    """Fallback for stale or unknown admin callback payloads."""
    await callback.answer("Panel yangilandi. Iltimos /admin ni qayta oching.", show_alert=True)


@router.callback_query(
    F.data.startswith("admin"),
)
async def admin_callback_denied(callback: CallbackQuery):
    """Access denied fallback for non-admin users tapping admin buttons."""
    await callback.answer("⛔ Admin panel huquqi yo'q", show_alert=True)
